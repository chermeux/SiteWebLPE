from datetime import datetime
from django.core.files.temp import NamedTemporaryFile

from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect

from GestionEleves.form import *
from GestionEleves.models import *
from Articles.views import getBase
from openpyxl import Workbook


# Create your views here.

@login_required
def liste(request):
    return render(request, 'GestionEleves/liste.html', getBase(request) |
                  {'eleves': Eleve.objects.all(),
                   'groupes': Groupe.objects.all(),
                   'appelles': Appelle.objects.all(),
                   'newEleveForm': EleveForm(),
                   'newGroupeForm': GroupeForm()})

@login_required
def newGroupe(request):
    form = GroupeForm(request.POST, request.FILES)
    form.save()
    return redirect('/GestionEleves/liste/')

@login_required
def newEleve(request):
    form = EleveForm(request.POST, request.FILES)
    form.save()
    return redirect('/GestionEleves/liste/')

@login_required
def newAppelle(request):
    appelle = Appelle(groupe_id=int(request.POST['groupeId']))
    appelle.save()

    for eleveId in [int(id) for id in request.POST.getlist('eleveId')]:
        Status(eleve_id=eleveId, appelle=appelle,
               present=str(eleveId) in request.POST and request.POST[str(eleveId)] == 'on').save()
    return redirect(f'/GestionEleves/groupe/{appelle.groupe.id}')

@login_required
def appelle(request, id):
    return render(request, 'GestionEleves/appel.html', getBase(request) | {'appelle': Appelle.objects.get(id=id)})

@login_required
def groupe(request, id):
    if request.method == 'POST':
        if request.POST['delete'] == 'je veux supprimer ce groupe':
            Groupe.objects.get(id=id).delete()
            return redirect('/GestionEleves/liste/')
    return render(request, 'GestionEleves/groupe.html', getBase(request) | {'groupe': Groupe.objects.get(id=id)})

@login_required
def eleve(request, id):
    if request.method == 'POST':
        if request.POST['delete'] == 'je veux supprimer cet élève':
            Eleve.objects.get(id=id).delete()
            return redirect('/GestionEleves/liste/')
    return render(request, 'GestionEleves/eleve.html', getBase(request) | {'eleve': Eleve.objects.get(id=id)})


def deleteAppelle(request, id):
    pass


def deleteGroupe(request, id):
    pass


def deleteEleve(request, id):
    pass


def getxl(request):
    wb = Workbook()
    for i, groupe in enumerate(Groupe.objects.all()):
        if i == 0:
            ws = wb.active
            ws.title = groupe.nom
        else:
            ws = wb.create_sheet(groupe.nom)
        ws.cell(row=1, column=1, value='Elève')
        for x, appel in enumerate(groupe.appelles):
            ws.cell(row=1, column=x + 2, value=appel.date.strftime('%Y-%m-%d'))
        for y, eleve in enumerate(groupe.eleves):
            ws.cell(row=y + 2, column=1, value=f'{eleve.prenom} {eleve.nom}')
            for x, appel in enumerate(eleve.presence):
                ws.cell(row=y + 2, column=x + 2, value=appel)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="Appel {datetime.now()}.xlsx"'},
    )
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response.write(tmp.read())
    return response
