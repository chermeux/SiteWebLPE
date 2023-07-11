from datetime import datetime
from django.core.files.temp import NamedTemporaryFile

from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect

from home.form import *
from home.models import *
from openpyxl import Workbook


# Create your views here.

def getBase(request):
	return {'matieres': Matiere.objects.all(), 'user': request.user}


def accueil(request):
	return render(request, "home/accueil.html", getBase(request))


def matieres(request, id):
	if id is None:
		cours = Cours.objects.filter(estCours=True)
	else:
		cours = Cours.objects.filter(matiere_id=id, estCours=True)
	if not request.user.is_authenticated:
		cours.filter(public=True)
	return render(request, "home/matieres.html", getBase(request) | {'cours': cours})


def orientation(request):
	cours = Cours.objects.filter(estCours=False)
	if not request.user.is_authenticated:
		cours.filter(public=True)
	return render(request, "home/orientation.html", getBase(request) | {'cours': cours})


def cordees(request):
	return render(request, "home/cordees.html", getBase(request))


def projets(request):
	return render(request, "home/projets.html", getBase(request))


def contact(request):
	return render(request, "home/contact.html", getBase(request))


def cours(request, id):
	return render(request, "home/cours.html", getBase(request) | {"cours": Cours.objects.get(id=id)})


@login_required
def coursAdmin(request, id):
	nv = False
	if request.method == 'POST':
		if request.POST['delete'] == 'je veux supprimer ce cours':
			Cours.objects.get(id=id).delete()
			return redirect('/matieres/')
		if id is not None:
			print(request.POST, request.FILES)
			form = CoursForm(request.POST, request.FILES, instance=Cours.objects.get(id=id))
			cours = form.save()
		else:
			form = CoursForm(request.POST, request.FILES)
			cours = form.save()
			id = cours.id
			nv = True
		files = request.FILES.getlist('file_field')
		for f in files:
			newFile = Fichier(nom=f.name[:40], fichier=f, cour_id=id)
			if newFile.fichier.name.endswith('.pdf'):
				newFile.save()
	if nv:
		return redirect(f'/ad/cours/{id}')
	return render(request, 'home/editCoursAdmin.html', getBase(request) |
	              {"coursForm": CoursForm(instance=Cours.objects.get(id=id)) if id is not None else CoursForm(),
	               'cours'    : Cours.objects.get(id=id) if id is not None else None,
	               'fileForm' : FileFieldForm(use_required_attribute=False)})


@login_required
def deleteFile(request, id):
	fichier = Fichier.objects.get(id=id)
	idCours = fichier.cour_id
	fichier.fichier.delete()
	fichier.delete()
	return redirect(f'/ad/cours/{idCours}')


def deleteCours(request, id):
	cours = Cours.objects.get(id=id)
	cours.delete()
	return redirect('/matieres/')


def loginPage(request):
	logout(request)
	if request.method == 'GET':
		if 'next' in request.GET:
			next = request.GET['next']
		else:
			next = '/'
		return render(request, 'home/login.html', {'next': next})
	user = authenticate(username=request.POST['user'], password=request.POST['password'])
	next = request.POST['next']
	if user is not None:
		login(request, user)
		return redirect(next)
	else:
		return render(request, 'home/login.html',
		              {'next': next, 'msg': 'Utilisateur ou mots de passe invalide', 'user': request.POST['user']})


@login_required
def liste(request):
	return render(request, 'home/liste.html', getBase(request) |
	              {'eleves'       : Eleve.objects.all(),
	               'groupes'      : Groupe.objects.all(),
	               'appelles'     : Appelle.objects.all(),
	               'newEleveForm' : EleveForm(),
	               'newGroupeForm': GroupeForm()})


def newGroupe(request):
	form = GroupeForm(request.POST, request.FILES)
	form.save()
	return redirect('/liste/')


def newEleve(request):
	form = EleveForm(request.POST, request.FILES)
	form.save()
	return redirect('/liste/')


def newAppelle(request):
	appelle = Appelle(groupe_id=int(request.POST['groupeId']))
	appelle.save()

	for eleveId in [int(id) for id in request.POST.getlist('eleveId')]:
		Status(eleve_id=eleveId, appelle=appelle,
		       present=str(eleveId) in request.POST and request.POST[str(eleveId)] == 'on').save()
	return redirect(f'/groupe/{appelle.groupe.id}')


def appelle(request, id):
	return render(request, 'home/appel.html', getBase(request) | {'appelle': Appelle.objects.get(id=id)})


def groupe(request, id):
	if request.method == 'POST':
		if request.POST['delete'] == 'je veux supprimer ce groupe':
			Groupe.objects.get(id=id).delete()
			return redirect('/liste/')
	return render(request, 'home/groupe.html', getBase(request) | {'groupe': Groupe.objects.get(id=id)})


def eleve(request, id):
	if request.method == 'POST':
		if request.POST['delete'] == 'je veux supprimer cet élève':
			Eleve.objects.get(id=id).delete()
			return redirect('/liste/')
	return render(request, 'home/eleve.html', getBase(request) | {'eleve': Eleve.objects.get(id=id)})


def deleteAppelle(request, id):
	pass


def deleteGroupe(request, id):
	pass


def deleteEleve(request, id):
	pass




def getxl(request):
	wb = Workbook()
	for i, groupe in enumerate(Groupe.objects.all()):
		if i == 0:
			ws = wb.active
			ws.title = groupe.nom
		else:
			ws = wb.create_sheet(groupe.nom)
		ws.cell(row=1, column=1, value='Elève')
		for x, appel in enumerate(groupe.appelles):
			ws.cell(row=1, column=x + 2, value=appel.date.strftime('%Y-%m-%d'))
		for y, eleve in enumerate(groupe.eleves):
			ws.cell(row=y + 2, column=1, value=f'{eleve.prenom} {eleve.nom}')
			for x, appel in enumerate(eleve.presence):
				ws.cell(row=y + 2, column=x + 2, value=appel)
	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		headers={'Content-Disposition': f'attachment; filename="Appel {datetime.now()}.xlsx"'},
	)
	with NamedTemporaryFile() as tmp:
		wb.save(tmp.name)
		tmp.seek(0)
		response.write(tmp.read())
	return response
